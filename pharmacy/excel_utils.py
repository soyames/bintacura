import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from django.http import HttpResponse
from django.db import models
from prescriptions.models import Medication
from pharmacy.models import PharmacyInventory, PharmacyStockMovement
from currency_converter.services import CurrencyConverterService


def generate_inventory_template():
    """Generate Excel template for inventory upload"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modèle Inventaire"
    
    # Header styling
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Nom du Médicament*',
        'Nom Générique',
        'Numéro de Lot*',
        'Quantité en Stock*',
        'Prix Unitaire (USD)*',
        'Prix de Vente (USD)*',
        'Fabricant',
        'Date de Fabrication',
        'Date d\'Expiration*',
        'Niveau de Réapprovisionnement',
        'Emplacement de Stockage',
        'Réfrigération Requise',
        'Disponible Publiquement'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Instructions row
    instructions = [
        'Paracétamol 500mg',
        'Paracétamol',
        'LOT001',
        '100',
        '500',
        '750',
        'ABC Pharma',
        '2024-01-15',
        '2026-01-15',
        '20',
        'A-01-01',
        'NON',
        'OUI'
    ]
    
    for col_num, value in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='left')
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.column_dimensions['A'].width = 50
    ws_instructions.column_dimensions['B'].width = 50
    
    instructions_data = [
        ("INSTRUCTIONS D'UTILISATION", ""),
        ("", ""),
        ("Champs Obligatoires (*)", "Description"),
        ("Nom du Médicament", "Nom commercial ou nom du médicament"),
        ("Numéro de Lot", "Numéro de lot du fabricant"),
        ("Quantité en Stock", "Nombre d'unités en stock (nombre entier)"),
        ("Prix Unitaire", "Prix d'achat par unité en USD"),
        ("Prix de Vente", "Prix de vente par unité en USD"),
        ("Date d'Expiration", "Format: AAAA-MM-JJ (ex: 2026-12-31)"),
        ("", ""),
        ("Champs Optionnels", "Description"),
        ("Nom Générique", "Nom générique du médicament"),
        ("Fabricant", "Nom du fabricant"),
        ("Date de Fabrication", "Format: AAAA-MM-JJ"),
        ("Niveau de Réapprovisionnement", "Quantité minimum avant alerte (défaut: 10)"),
        ("Emplacement de Stockage", "Code d'emplacement dans la pharmacie"),
        ("Réfrigération Requise", "OUI ou NON (défaut: NON)"),
        ("Disponible Publiquement", "OUI ou NON (défaut: OUI)"),
        ("", ""),
        ("NOTES IMPORTANTES", ""),
        ("1. La première ligne est l'exemple", "Ne pas supprimer les en-têtes"),
        ("2. Les dates doivent être au format AAAA-MM-JJ", ""),
        ("3. Les prix sont en USD (ex: 1.50 pour $1.50)", ""),
        ("4. OUI/NON pour les champs booléens", ""),
        ("5. Si le médicament n'existe pas, il sera créé", ""),
    ]
    
    for row_num, (col1, col2) in enumerate(instructions_data, 1):
        ws_instructions.cell(row=row_num, column=1, value=col1)
        ws_instructions.cell(row=row_num, column=2, value=col2)
        if row_num == 1:
            ws_instructions.cell(row=row_num, column=1).font = Font(bold=True, size=14, color="10b981")
        elif row_num in [3, 11, 20]:
            ws_instructions.cell(row=row_num, column=1).font = Font(bold=True, size=12)
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Generate file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def import_inventory_from_excel(file, pharmacy):
    """Import inventory data from Excel file with currency conversion"""
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # Get pharmacy's local currency for displaying converted prices
        pharmacy_currency = CurrencyConverterService.get_participant_currency(pharmacy)
        
        results = {
            'success': 0,
            'errors': [],
            'warnings': [],
            'created_medications': []
        }
        
        # Skip header row
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            try:
                # Extract data
                med_name = row[0]
                generic_name = row[1] or ''
                batch_number = row[2]
                quantity = row[3]
                unit_price = row[4]
                selling_price = row[5]
                manufacturer = row[6] or ''
                manufacturing_date = row[7]
                expiry_date = row[8]
                reorder_level = row[9] or 10
                storage_location = row[10] or ''
                requires_refrigeration = str(row[11]).upper() in ['OUI', 'YES', 'TRUE', '1'] if row[11] else False
                is_publicly_available = str(row[12]).upper() in ['OUI', 'YES', 'TRUE', '1'] if row[12] is not None else True
                
                # Validate required fields
                if not med_name:
                    results['errors'].append(f"Ligne {row_num}: Nom du médicament manquant")
                    continue
                if not batch_number:
                    results['errors'].append(f"Ligne {row_num}: Numéro de lot manquant")
                    continue
                if not quantity:
                    results['errors'].append(f"Ligne {row_num}: Quantité manquante")
                    continue
                if not unit_price:
                    results['errors'].append(f"Ligne {row_num}: Prix unitaire manquant")
                    continue
                if not selling_price:
                    results['errors'].append(f"Ligne {row_num}: Prix de vente manquant")
                    continue
                if not expiry_date:
                    results['errors'].append(f"Ligne {row_num}: Date d'expiration manquante")
                    continue
                
                # Convert quantities to integers
                try:
                    quantity = int(float(quantity))
                    # Prices are in USD, convert to minor units (USD cents)
                    unit_price = int(float(unit_price) * 100)
                    selling_price = int(float(selling_price) * 100)
                    reorder_level = int(float(reorder_level))
                except ValueError:
                    results['errors'].append(f"Ligne {row_num}: Format numérique invalide")
                    continue
                
                # Parse dates
                if isinstance(expiry_date, datetime):
                    expiry_date = expiry_date.date()
                elif isinstance(expiry_date, str):
                    try:
                        expiry_date = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                    except ValueError:
                        results['errors'].append(f"Ligne {row_num}: Format de date d'expiration invalide (utilisez AAAA-MM-JJ)")
                        continue
                
                if manufacturing_date:
                    if isinstance(manufacturing_date, datetime):
                        manufacturing_date = manufacturing_date.date()
                    elif isinstance(manufacturing_date, str):
                        try:
                            manufacturing_date = datetime.strptime(manufacturing_date, '%Y-%m-%d').date()
                        except ValueError:
                            results['warnings'].append(f"Ligne {row_num}: Format de date de fabrication invalide, ignoré")
                            manufacturing_date = None
                
                # Find or create medication
                medication = Medication.objects.filter(name__iexact=med_name).first()
                if not medication:
                    medication = Medication.objects.create(
                        name=med_name,
                        generic_name=generic_name,
                        manufacturer=manufacturer,
                        category='Imported',
                        requires_prescription=True
                    )
                    results['created_medications'].append(med_name)
                
                # Check if inventory item already exists
                existing_item = PharmacyInventory.objects.filter(
                    pharmacy=pharmacy,
                    medication=medication,
                    batch_number=batch_number
                ).first()
                
                if existing_item:
                    # Update existing item
                    existing_item.quantity_in_stock = quantity
                    existing_item.unit_price = unit_price
                    existing_item.selling_price = selling_price
                    existing_item.manufacturer = manufacturer
                    existing_item.manufacturing_date = manufacturing_date
                    existing_item.expiry_date = expiry_date
                    existing_item.reorder_level = reorder_level
                    existing_item.storage_location = storage_location
                    existing_item.requires_refrigeration = requires_refrigeration
                    existing_item.is_publicly_available = is_publicly_available
                    existing_item.save()
                    
                    results['warnings'].append(f"Ligne {row_num}: Article existant mis à jour - {med_name} (Lot: {batch_number})")
                else:
                    # Create new inventory item
                    PharmacyInventory.objects.create(
                        pharmacy=pharmacy,
                        medication=medication,
                        batch_number=batch_number,
                        quantity_in_stock=quantity,
                        unit_price=unit_price,
                        selling_price=selling_price,
                        manufacturer=manufacturer,
                        manufacturing_date=manufacturing_date,
                        expiry_date=expiry_date,
                        reorder_level=reorder_level,
                        storage_location=storage_location,
                        requires_refrigeration=requires_refrigeration,
                        is_publicly_available=is_publicly_available
                    )
                
                results['success'] += 1
                
            except Exception as e:
                results['errors'].append(f"Ligne {row_num}: Erreur - {str(e)}")
        
        return results
        
    except Exception as e:
        return {
            'success': 0,
            'errors': [f"Erreur de lecture du fichier: {str(e)}"],
            'warnings': [],
            'created_medications': []
        }


def export_inventory_to_excel(pharmacy):
    """Export current inventory to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventaire"
    
    # Header styling
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Médicament',
        'Nom Générique',
        'Numéro de Lot',
        'Quantité',
        'Prix Unitaire',
        'Prix de Vente',
        'Valeur Totale',
        'Fabricant',
        'Date Fabrication',
        'Date Expiration',
        'Jours Restants',
        'Niveau Réapprovisionnement',
        'Emplacement',
        'Réfrigération',
        'Statut'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Get inventory items
    items = PharmacyInventory.objects.filter(pharmacy=pharmacy).select_related('medication')
    
    # Data rows
    for row_num, item in enumerate(items, start=2):
        days_to_expiry = (item.expiry_date - datetime.now().date()).days
        
        if item.quantity_in_stock == 0:
            status = 'Rupture'
        elif item.quantity_in_stock <= item.reorder_level:
            status = 'Stock Faible'
        else:
            status = 'En Stock'
        
        data = [
            item.medication.name,
            item.medication.generic_name,
            item.batch_number,
            item.quantity_in_stock,
            item.unit_price,
            item.selling_price,
            item.quantity_in_stock * item.unit_price,
            item.manufacturer,
            item.manufacturing_date.strftime('%Y-%m-%d') if item.manufacturing_date else '',
            item.expiry_date.strftime('%Y-%m-%d'),
            days_to_expiry,
            item.reorder_level,
            item.storage_location,
            'OUI' if item.requires_refrigeration else 'NON',
            status
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Color code based on status
            if col_num == 15:  # Status column
                if status == 'Rupture':
                    cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
                elif status == 'Stock Faible':
                    cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
            
            # Color code expiry
            if col_num == 11:  # Days remaining
                if days_to_expiry < 30:
                    cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
                elif days_to_expiry < 90:
                    cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    # Generate file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
