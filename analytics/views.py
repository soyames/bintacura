from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.views import APIView
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from datetime import timedelta, datetime
from decimal import Decimal

from .models import PlatformStatistics, UserGrowthMetrics, RevenueMetrics, SurveyResponse
from .serializers import (
    PlatformStatisticsSerializer,
    UserGrowthMetricsSerializer,
    RevenueMetricsSerializer,
    DashboardOverviewSerializer,
    UserGrowthDataSerializer,
    RevenueDataSerializer,
    RoleDistributionSerializer,
    ActivitySerializer,
    TopProviderSerializer,
    GeographicDistributionSerializer,
    DoctorAnalyticsSerializer,
    HospitalAnalyticsSerializer,
    PharmacyAnalyticsSerializer,
    InsuranceAnalyticsSerializer,
    PatientAnalyticsSerializer,
    SurveyStatisticsSerializer,
)
from .services import AnalyticsService
from core.models import Participant, Wallet, Transaction
from appointments.models import Appointment
from prescriptions.models import Prescription
from insurance.models import InsuranceClaim, InsuranceSubscription


class PlatformStatisticsViewSet(viewsets.ReadOnlyModelViewSet):  # View for PlatformStatisticsSet operations
    queryset = PlatformStatistics.objects.all()
    serializer_class = PlatformStatisticsSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):  # Get queryset
        if not (
            self.request.user.is_superuser or self.request.user.role == "super_admin"
        ):
            return PlatformStatistics.objects.none()
        return super().get_queryset()


class AdminAnalyticsView(APIView):  # Admin configuration for AnalyticsView model
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if not (request.user.is_superuser or request.user.role == "super_admin"):
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        overview = AnalyticsService.get_dashboard_overview()
        serializer = DashboardOverviewSerializer(overview)
        return Response(serializer.data)


class ParticipantGrowthView(APIView):  # View for ParticipantGrowth operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if not (request.user.is_superuser or request.user.role == "super_admin"):
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        days = int(request.query_params.get("days", 30))
        growth_data = AnalyticsService.get_user_growth_data(days=days)
        serializer = UserGrowthDataSerializer(growth_data, many=True)
        return Response(serializer.data)


class RevenueAnalyticsView(APIView):  # View for RevenueAnalytics operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if not (request.user.is_superuser or request.user.role == "super_admin"):
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        days = int(request.query_params.get("days", 30))
        revenue_data = AnalyticsService.get_revenue_data(days=days)
        serializer = RevenueDataSerializer(revenue_data, many=True)
        return Response(serializer.data)


class RoleDistributionView(APIView):  # View for RoleDistribution operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if not (request.user.is_superuser or request.user.role == "super_admin"):
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        distribution = AnalyticsService.get_role_distribution()
        serializer = RoleDistributionSerializer(distribution, many=True)
        return Response(serializer.data)


class RecentActivitiesView(APIView):  # View for RecentActivities operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if not (request.user.is_superuser or request.user.role == "super_admin"):
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        limit = int(request.query_params.get("limit", 20))
        activities = AnalyticsService.get_recent_activities(limit=limit)
        serializer = ActivitySerializer(activities, many=True)
        return Response(serializer.data)


class TopProvidersView(APIView):  # View for TopProviders operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if not (request.user.is_superuser or request.user.role == "super_admin"):
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        limit = int(request.query_params.get("limit", 10))
        providers = AnalyticsService.get_top_providers(limit=limit)
        serializer = TopProviderSerializer(providers, many=True)
        return Response(serializer.data)


class GeographicDistributionView(APIView):  # View for GeographicDistribution operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if not (request.user.is_superuser or request.user.role == "super_admin"):
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        distribution = AnalyticsService.get_geographic_distribution()
        serializer = GeographicDistributionSerializer(distribution, many=True)
        return Response(serializer.data)


class DoctorAnalyticsView(APIView):  # View for DoctorAnalytics operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if request.user.role != "doctor":
            return Response(
                {"error": "Only doctors can access this endpoint"},
                status=status.HTTP_403_FORBIDDEN,
            )

        doctor = request.user

        appointments = Appointment.objects.filter(doctor=doctor)
        total_appointments = appointments.count()
        completed = appointments.filter(status="completed").count()
        cancelled = appointments.filter(status="cancelled").count()
        pending = appointments.filter(status__in=["scheduled", "confirmed"]).count()

        transactions = Transaction.objects.filter(sender=doctor, status="completed")
        total_revenue = transactions.aggregate(total=Sum("amount"))["total"] or Decimal(
            "0"
        )

        patients = appointments.values("patient").distinct().count()

        completion_rate = (
            (completed / total_appointments * 100) if total_appointments > 0 else 0
        )

        data = {
            "total_appointments": total_appointments,
            "completed_appointments": completed,
            "cancelled_appointments": cancelled,
            "pending_appointments": pending,
            "total_revenue": float(total_revenue),
            "total_patients": patients,
            "average_rating": 0.0,
            "completion_rate": round(completion_rate, 2),
        }

        serializer = DoctorAnalyticsSerializer(data)
        return Response(serializer.data)


class HospitalAnalyticsView(APIView):  # View for HospitalAnalytics operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if request.user.role != "hospital":
            return Response(
                {"error": "Only hospitals can access this endpoint"},
                status=status.HTTP_403_FORBIDDEN,
            )

        hospital = request.user

        total_doctors = Participant.objects.filter(
            role="doctor", provider_data__hospital=hospital, is_active=True
        ).count()

        appointments = Appointment.objects.filter(
            Q(doctor__provider_data__hospital=hospital) | Q(hospital=hospital)
        )
        total_appointments = appointments.count()

        patients = appointments.values("patient").distinct().count()

        transactions = Transaction.objects.filter(sender=hospital, status="completed")
        total_revenue = transactions.aggregate(total=Sum("amount"))["total"] or Decimal(
            "0"
        )

        data = {
            "total_doctors": total_doctors,
            "total_appointments": total_appointments,
            "total_patients": patients,
            "total_revenue": float(total_revenue),
            "bed_occupancy_rate": 0.0,
            "average_rating": 0.0,
        }

        serializer = HospitalAnalyticsSerializer(data)
        return Response(serializer.data)


class PharmacyAnalyticsView(APIView):  # View for PharmacyAnalytics operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if request.user.role != "pharmacy":
            return Response(
                {"error": "Only pharmacies can access this endpoint"},
                status=status.HTTP_403_FORBIDDEN,
            )

        pharmacy = request.user

        prescriptions = Prescription.objects.filter(pharmacy=pharmacy)
        total_prescriptions = prescriptions.count()
        fulfilled = prescriptions.filter(status="fulfilled").count()
        pending = prescriptions.filter(status="pending").count()

        transactions = Transaction.objects.filter(sender=pharmacy, status="completed")
        total_revenue = transactions.aggregate(total=Sum("amount"))["total"] or Decimal(
            "0"
        )

        customers = prescriptions.values("patient").distinct().count()

        avg_order = (
            (total_revenue / total_prescriptions)
            if total_prescriptions > 0
            else Decimal("0")
        )
        fulfillment_rate = (
            (fulfilled / total_prescriptions * 100) if total_prescriptions > 0 else 0
        )

        data = {
            "total_prescriptions": total_prescriptions,
            "fulfilled_prescriptions": fulfilled,
            "pending_prescriptions": pending,
            "total_revenue": float(total_revenue),
            "total_customers": customers,
            "average_order_value": float(avg_order),
            "fulfillment_rate": round(fulfillment_rate, 2),
        }

        serializer = PharmacyAnalyticsSerializer(data)
        return Response(serializer.data)


class InsuranceAnalyticsView(APIView):  # View for InsuranceAnalytics operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if request.user.role != "insurance_company":
            return Response(
                {"error": "Only insurance companies can access this endpoint"},
                status=status.HTTP_403_FORBIDDEN,
            )

        insurance = request.user

        claims = InsuranceClaim.objects.filter(insurance_company=insurance)
        total_claims = claims.count()
        approved = claims.filter(status="approved").count()
        pending = claims.filter(status="pending").count()
        rejected = claims.filter(status="rejected").count()

        total_claims_amount = claims.aggregate(total=Sum("claim_amount"))[
            "total"
        ] or Decimal("0")

        total_approved_amount = claims.filter(status="approved").aggregate(
            total=Sum("approved_amount")
        )["total"] or Decimal("0")

        approval_rate = (approved / total_claims * 100) if total_claims > 0 else 0

        active_subs = InsuranceSubscription.objects.filter(
            insurance_package__insurance_company=insurance, status="active"
        ).count()

        data = {
            "total_claims": total_claims,
            "approved_claims": approved,
            "pending_claims": pending,
            "rejected_claims": rejected,
            "total_claims_amount": float(total_claims_amount),
            "total_approved_amount": float(total_approved_amount),
            "approval_rate": round(approval_rate, 2),
            "active_subscriptions": active_subs,
        }

        serializer = InsuranceAnalyticsSerializer(data)
        return Response(serializer.data)


class PatientAnalyticsView(APIView):  # View for PatientAnalytics operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        if request.user.role != "patient":
            return Response(
                {"error": "Only patients can access this endpoint"},
                status=status.HTTP_403_FORBIDDEN,
            )

        patient = request.user

        appointments = Appointment.objects.filter(patient=patient)
        total_appointments = appointments.count()
        completed = appointments.filter(status="completed").count()

        prescriptions = Prescription.objects.filter(patient=patient).count()

        transactions = Transaction.objects.filter(sender=patient, status="completed")
        total_spending = transactions.aggregate(total=Sum("amount"))[
            "total"
        ] or Decimal("0")

        wallet = Wallet.objects.filter(participant=patient).first()
        wallet_balance = wallet.balance if wallet else Decimal("0")

        claims = InsuranceClaim.objects.filter(patient=patient).count()

        data = {
            "total_appointments": total_appointments,
            "completed_appointments": completed,
            "total_prescriptions": prescriptions,
            "total_spending": float(total_spending),
            "wallet_balance": float(wallet_balance),
            "insurance_claims": claims,
        }

        serializer = PatientAnalyticsSerializer(data)
        return Response(serializer.data)


class DetailedPatientAnalyticsView(APIView):  # View for DetailedPatientAnalytics operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        from .analytics_service import PatientAnalytics

        if request.user.role != "patient":
            return Response(
                {"error": "Only patients can access this endpoint"},
                status=status.HTTP_403_FORBIDDEN,
            )

        stats = PatientAnalytics.get_dashboard_stats(request.user)
        return Response(stats)


class DetailedDoctorAnalyticsView(APIView):  # View for DetailedDoctorAnalytics operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        from .analytics_service import DoctorAnalytics

        if request.user.role != "doctor":
            return Response(
                {"error": "Only doctors can access this endpoint"},
                status=status.HTTP_403_FORBIDDEN,
            )

        stats = DoctorAnalytics.get_dashboard_stats(request.user)
        return Response(stats)


class DetailedHospitalAnalyticsView(APIView):  # View for DetailedHospitalAnalytics operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        from .analytics_service import HospitalAnalytics

        if request.user.role != "hospital":
            return Response(
                {"error": "Only hospitals can access this endpoint"},
                status=status.HTTP_403_FORBIDDEN,
            )

        stats = HospitalAnalytics.get_dashboard_stats(request.user)
        return Response(stats)


class DetailedPharmacyAnalyticsView(APIView):  # View for DetailedPharmacyAnalytics operations
    permission_classes = [IsAuthenticated]

    def get(self, request):  # Get
        from .analytics_service import PharmacyAnalytics

        if request.user.role != "pharmacy":
            return Response(
                {"error": "Only pharmacies can access this endpoint"},
                status=status.HTTP_403_FORBIDDEN,
            )

        stats = PharmacyAnalytics.get_dashboard_stats(request.user)
        return Response(stats)


def survey_stats_view(request):  # Survey stats view
    return render(request, 'analytics/survey_stats.html')


def survey_submit_view(request):  # Survey submit view
    from django.core.cache import cache
    from django.contrib import messages

    if request.method == 'POST':
        ip_address = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', '')).split(',')[0].strip()

        rate_limit_key = f'survey_submit_{ip_address}'
        submissions_count = cache.get(rate_limit_key, 0)

        if submissions_count >= 3:
            messages.error(request, 'Trop de tentatives. Veuillez réessayer dans 1 heure.')
            return render(request, 'analytics/survey_form.html', {
                'form': None,
                'rate_limited': True
            })

        from .forms import SurveyResponseForm
        form = SurveyResponseForm(request.POST)
        if form.is_valid():
            form.save()

            cache.set(rate_limit_key, submissions_count + 1, 3600)

            return redirect('analytics_survey:survey_thank_you')
        else:
            return render(request, 'analytics/survey_form.html', {'form': form})
    else:
        from .forms import SurveyResponseForm
        form = SurveyResponseForm()
        return render(request, 'analytics/survey_form.html', {'form': form})


def survey_thank_you_view(request):  # Survey thank you view
    return render(request, 'analytics/survey_thank_you.html')


class SurveyStatisticsAPIView(APIView):  # View for SurveyStatisticsAPI operations
    permission_classes = [AllowAny]

    def get(self, request):  # Get
        from django.db.models import Count, Avg, Min, Max, StdDev

        responses = SurveyResponse.objects.all()
        total_responses = responses.count()

        if total_responses == 0:
            return Response({
                'total_responses': 0,
                'sex_distribution': {},
                'profession_distribution': {},
                'country_distribution': {},
                'price_stats': {},
                'currency_distribution': {},
                'recent_responses': []
            })

        sex_dist = dict(responses.values('sex').annotate(count=Count('sex')).values_list('sex', 'count'))

        profession_dist = dict(
            responses.values('profession')
            .annotate(count=Count('profession'))
            .order_by('-count')[:15]
            .values_list('profession', 'count')
        )

        country_dist = dict(
            responses.values('country')
            .annotate(count=Count('country'))
            .order_by('-count')[:15]
            .values_list('country', 'count')
        )

        currency_dist = dict(
            responses.values('currency')
            .annotate(count=Count('currency'))
            .values_list('currency', 'count')
        )

        price_responses = responses.exclude(suggested_price__isnull=True)
        price_stats = {}
        if price_responses.exists():
            price_agg = price_responses.aggregate(
                avg=Avg('suggested_price'),
                min=Min('suggested_price'),
                max=Max('suggested_price'),
                std=StdDev('suggested_price')
            )
            price_stats = {
                'average': float(price_agg['avg']) if price_agg['avg'] else 0,
                'min': float(price_agg['min']) if price_agg['min'] else 0,
                'max': float(price_agg['max']) if price_agg['max'] else 0,
                'std_dev': float(price_agg['std']) if price_agg['std'] else 0,
            }

            prices_list = list(price_responses.values_list('suggested_price', flat=True))
            price_stats['values'] = [float(p) for p in prices_list]

        recent = list(
            responses.order_by('-submission_date')[:10]
            .values('email', 'country', 'profession', 'sex', 'submission_date')
        )
        for item in recent:
            item['submission_date'] = item['submission_date'].isoformat()

        data = {
            'total_responses': total_responses,
            'sex_distribution': sex_dist,
            'profession_distribution': profession_dist,
            'country_distribution': country_dist,
            'price_stats': price_stats,
            'currency_distribution': currency_dist,
            'recent_responses': recent
        }

        serializer = SurveyStatisticsSerializer(data)
        return Response(serializer.data)


@staff_member_required
def export_survey_data_view(request, file_format):  # Export survey data view
    if not request.user.is_superuser:
        return HttpResponse("Access Denied: Superuser status required.", status=403)

    try:
        responses = SurveyResponse.objects.all()

        if not responses.exists():
            return HttpResponse("No survey data available to export.", status=404)

        if file_format == 'csv':
            import csv
            response = HttpResponse(content_type='text/csv')
            filename = f'survey_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            writer = csv.writer(response)
            writer.writerow(['ID', 'Email', 'Country', 'City', 'Profession', 'Sex', 'Suggested Price', 'Currency', 'Feature Suggestion', 'Other Suggestion', 'Submission Date'])

            for obj in responses:
                writer.writerow([
                    str(obj.id),
                    obj.email,
                    obj.country,
                    obj.city,
                    obj.profession,
                    obj.sex,
                    str(obj.suggested_price) if obj.suggested_price else '',
                    obj.currency,
                    obj.feature_suggestion,
                    obj.other_suggestion,
                    obj.submission_date.strftime('%Y-%m-%d %H:%M:%S')
                ])

        elif file_format == 'excel':
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = "Survey Responses"

            headers = ['ID', 'Email', 'Country', 'City', 'Profession', 'Sex', 'Suggested Price', 'Currency', 'Feature Suggestion', 'Other Suggestion', 'Submission Date']
            ws.append(headers)

            header_fill = PatternFill(start_color='4a90e2', end_color='4a90e2', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for obj in responses:
                ws.append([
                    str(obj.id),
                    obj.email,
                    obj.country,
                    obj.city,
                    obj.profession,
                    obj.sex,
                    float(obj.suggested_price) if obj.suggested_price else None,
                    obj.currency,
                    obj.feature_suggestion,
                    obj.other_suggestion,
                    obj.submission_date
                ])

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'survey_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

        elif file_format == 'pdf':
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from django.db.models import Count, Avg
            import io

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
            elements = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#4a90e2'),
                spaceAfter=30,
                alignment=1
            )

            elements.append(Paragraph("BINTACURA Survey Report", title_style))
            elements.append(Spacer(1, 0.3*inch))

            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=12
            )

            elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))

            total_responses = responses.count()
            elements.append(Paragraph(f"Total Responses: {total_responses}", subtitle_style))
            elements.append(Spacer(1, 0.2*inch))

            sex_dist = responses.values('sex').annotate(count=Count('sex'))
            sex_data = [['Sex', 'Count']]
            for item in sex_dist:
                sex_data.append([item['sex'], item['count']])

            sex_table = Table(sex_data)
            sex_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(Paragraph("Sex Distribution", subtitle_style))
            elements.append(sex_table)
            elements.append(Spacer(1, 0.3*inch))

            country_dist = responses.values('country').annotate(count=Count('country')).order_by('-count')[:10]
            country_data = [['Country', 'Count']]
            for item in country_dist:
                country_data.append([item['country'], item['count']])

            country_table = Table(country_data)
            country_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(Paragraph("Top 10 Countries", subtitle_style))
            elements.append(country_table)
            elements.append(Spacer(1, 0.3*inch))

            price_agg = responses.exclude(suggested_price__isnull=True).aggregate(
                avg=Avg('suggested_price')
            )
            if price_agg['avg']:
                elements.append(Paragraph(f"Average Suggested Price: ${price_agg['avg']:.2f}", subtitle_style))

            doc.build(elements)
            buffer.seek(0)

            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            filename = f'survey_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

        else:
            return HttpResponse("Invalid file format specified.", status=400)

        return response

    except Exception as e:
        return HttpResponse(f"Error generating export: {str(e)}", status=500)


class PredictiveAnalyticsViewSet(viewsets.ViewSet):
    """
    AI-Powered Predictive Analytics ViewSet
    Provides statistical forecasting and pattern analysis
    """
    permission_classes = [IsAuthenticated]

    def _check_admin_permission(self, request):
        """Check if user has admin permissions"""
        if not (request.user.is_superuser or request.user.role == 'super_admin'):
            return Response(
                {'error': 'Permission denied. Admin access required.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return None

    @action(detail=False, methods=['get'])
    def user_growth_forecast(self, request):
        """
        GET /api/analytics/predictive/user_growth_forecast/?days_forward=30&historical_days=90

        Predicts user growth using linear regression and trend analysis
        """
        permission_error = self._check_admin_permission(request)
        if permission_error:
            return permission_error

        from .predictive_analytics import PredictiveAnalytics

        days_forward = int(request.query_params.get('days_forward', 30))
        historical_days = int(request.query_params.get('historical_days', 90))

        try:
            forecast = PredictiveAnalytics.predict_user_growth(
                days_forward=days_forward,
                historical_days=historical_days
            )
            return Response(forecast, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {
                    'error': 'An error occurred while generating user growth forecast',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def revenue_forecast(self, request):
        """
        GET /api/analytics/predictive/revenue_forecast/?days_forward=30&historical_days=90

        Forecasts revenue using moving averages and trend detection
        """
        permission_error = self._check_admin_permission(request)
        if permission_error:
            return permission_error

        from .predictive_analytics import PredictiveAnalytics

        days_forward = int(request.query_params.get('days_forward', 30))
        historical_days = int(request.query_params.get('historical_days', 90))

        try:
            forecast = PredictiveAnalytics.forecast_revenue(
                days_forward=days_forward,
                historical_days=historical_days
            )
            return Response(forecast, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {
                    'error': 'An error occurred while generating revenue forecast',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def appointment_completion_prediction(self, request):
        """
        GET /api/analytics/predictive/appointment_completion_prediction/?days_forward=30

        Predicts appointment completion rates based on historical patterns
        """
        permission_error = self._check_admin_permission(request)
        if permission_error:
            return permission_error

        from .predictive_analytics import PredictiveAnalytics

        days_forward = int(request.query_params.get('days_forward', 30))

        try:
            prediction = PredictiveAnalytics.predict_appointment_completion_rate(
                days_forward=days_forward
            )
            return Response(prediction, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {
                    'error': 'An error occurred while predicting appointment completion',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def platform_usage_analysis(self, request):
        """
        GET /api/analytics/predictive/platform_usage_analysis/?days=90

        Analyzes platform usage patterns and engagement metrics
        """
        permission_error = self._check_admin_permission(request)
        if permission_error:
            return permission_error

        from .predictive_analytics import PredictiveAnalytics

        days = int(request.query_params.get('days', 90))

        try:
            analysis = PredictiveAnalytics.analyze_platform_usage_patterns(days=days)
            return Response(analysis, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {
                    'error': 'An error occurred while analyzing platform usage',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def predictive_insights_overview(self, request):
        """
        GET /api/analytics/predictive/predictive_insights_overview/

        Aggregates all predictive analytics insights
        """
        permission_error = self._check_admin_permission(request)
        if permission_error:
            return permission_error

        from .predictive_analytics import PredictiveAnalytics

        try:
            insights = PredictiveAnalytics.get_predictive_insights()
            return Response({
                'insights': insights,
                'total_insights': len(insights),
                'generated_at': timezone.now()
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {
                    'error': 'An error occurred while generating predictive insights',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def ml_patient_segmentation(self, request):
        """
        GET /api/analytics/predictive/ml_patient_segmentation/?n_clusters=4

        ML-powered patient segmentation using K-Means clustering (Phase 7)
        """
        permission_error = self._check_admin_permission(request)
        if permission_error:
            return permission_error

        try:
            from ml_models.patient_segmentation import PatientSegmentation

            n_clusters = int(request.query_params.get('n_clusters', 4))
            if n_clusters < 2 or n_clusters > 10:
                return Response(
                    {'error': 'n_clusters must be between 2 and 10'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Segment patients for requesting organization
            user = request.user
            segmentation_result = PatientSegmentation.segment_patients(user, n_clusters=n_clusters)

            return Response({
                'organization': user.full_name,
                'ml_model': 'K-Means Clustering',
                'segmentation': segmentation_result,
                'generated_at': timezone.now()
            }, status=status.HTTP_200_OK)

        except ImportError:
            return Response(
                {'error': 'ML models not available. Install scikit-learn.'},
                status=status.HTTP_501_NOT_IMPLEMENTED
            )
        except Exception as e:
            return Response(
                {
                    'error': 'Patient segmentation failed',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def ml_advanced_revenue_forecast(self, request):
        """
        GET /api/analytics/predictive/ml_advanced_revenue_forecast/?days_forward=30&historical_days=180

        Advanced ML revenue forecast with seasonality (Phase 7)
        """
        permission_error = self._check_admin_permission(request)
        if permission_error:
            return permission_error

        try:
            from ml_models.revenue_forecast import AdvancedRevenueForecast

            days_forward = int(request.query_params.get('days_forward', 30))
            historical_days = int(request.query_params.get('historical_days', 180))

            if days_forward < 1 or days_forward > 365:
                return Response(
                    {'error': 'days_forward must be between 1 and 365'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if historical_days < 30 or historical_days > 730:
                return Response(
                    {'error': 'historical_days must be between 30 and 730'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get ML forecast for requesting organization
            user = request.user
            forecast = AdvancedRevenueForecast.forecast_revenue(user, days_forward, historical_days)

            return Response({
                'organization': user.full_name,
                'ml_model': 'Linear Regression with Seasonality',
                'forecast': forecast,
                'generated_at': timezone.now()
            }, status=status.HTTP_200_OK)

        except ImportError:
            return Response(
                {'error': 'ML models not available. Install scikit-learn.'},
                status=status.HTTP_501_NOT_IMPLEMENTED
            )
        except Exception as e:
            return Response(
                {
                    'error': 'Revenue forecast failed',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def ml_revenue_comparison(self, request):
        """
        GET /api/analytics/predictive/ml_revenue_comparison/?days_forward=30

        Compare ML forecast vs baseline moving average (Phase 7)
        """
        permission_error = self._check_admin_permission(request)
        if permission_error:
            return permission_error

        try:
            from ml_models.revenue_forecast import AdvancedRevenueForecast

            days_forward = int(request.query_params.get('days_forward', 30))

            if days_forward < 1 or days_forward > 365:
                return Response(
                    {'error': 'days_forward must be between 1 and 365'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get comparison for requesting organization
            user = request.user
            comparison = AdvancedRevenueForecast.compare_with_baseline(user, days_forward)

            return Response({
                'organization': user.full_name,
                'comparison': comparison,
                'generated_at': timezone.now()
            }, status=status.HTTP_200_OK)

        except ImportError:
            return Response(
                {'error': 'ML models not available. Install scikit-learn.'},
                status=status.HTTP_501_NOT_IMPLEMENTED
            )
        except Exception as e:
            return Response(
                {
                    'error': 'Revenue comparison failed',
                    'detail': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

